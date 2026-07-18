import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Python MCQ Quiz"

questions = [
    {
        "q": "What is the correct file extension for Python files?",
        "a": ".py",
        "b": ".python",
        "c": ".pt",
        "d": ".pyt",
        "ans": "A"
    },
    {
        "q": "Which keyword is used to define a function in Python?",
        "a": "func",
        "b": "function",
        "c": "def",
        "d": "define",
        "ans": "C"
    },
    {
        "q": "What is the output of: print(type(10))?",
        "a": "<class 'float'>",
        "b": "<class 'int'>",
        "c": "<class 'str'>",
        "d": "<class 'number'>",
        "ans": "B"
    },
    {
        "q": "Which of the following is used to create a comment in Python?",
        "a": "//",
        "b": "#",
        "c": "/* */",
        "d": "--",
        "ans": "B"
    },
    {
        "q": "What is the correct way to create a variable in Python?",
        "a": "x := 10",
        "b": "var x = 10",
        "c": "x = 10",
        "d": "int x = 10",
        "ans": "C"
    },
    {
        "q": "What is the output of: print(2 ** 3)?",
        "a": "6",
        "b": "8",
        "c": "5",
        "d": "23",
        "ans": "B"
    },
    {
        "q": "Which data type is mutable in Python?",
        "a": "Tuple",
        "b": "String",
        "c": "List",
        "d": "Integer",
        "ans": "C"
    },
    {
        "q": "What does len() function do?",
        "a": "Converts to lowercase",
        "b": "Returns the length of an object",
        "c": "Returns the last element",
        "d": "Deletes an element",
        "ans": "B"
    },
    {
        "q": "Which of the following is a valid Python loop?",
        "a": "for i in range(10):",
        "b": "for (i=0; i<10; i++):",
        "c": "for i = 1 to 10:",
        "d": "foreach i in 10:",
        "ans": "A"
    },
    {
        "q": "What is the output of: print('Hello' + ' World')?",
        "a": "Hello World",
        "b": "HelloWorld",
        "c": "Hello + World",
        "d": "Error",
        "ans": "A"
    },
    {
        "q": "Which method is used to add an element to the end of a list?",
        "a": "add()",
        "b": "append()",
        "c": "insert()",
        "d": "push()",
        "ans": "B"
    },
    {
        "q": "What is the output of: print(isinstance(5, int))?",
        "a": "False",
        "b": "True",
        "c": "5",
        "d": "Error",
        "ans": "B"
    },
    {
        "q": "What is the output of: print(10 // 3)?",
        "a": "3.33",
        "b": "3",
        "c": "4",
        "d": "3.0",
        "ans": "B"
    },
    {
        "q": "Which keyword is used for conditional statements in Python?",
        "a": "when",
        "b": "switch",
        "c": "if",
        "d": "condition",
        "ans": "C"
    },
    {
        "q": "What is the output of: print(bool(''))?",
        "a": "True",
        "b": "False",
        "c": "None",
        "d": "0",
        "ans": "B"
    },
    {
        "q": "How do you create a list in Python?",
        "a": "list = (1, 2, 3)",
        "b": "list = [1, 2, 3]",
        "c": "list = {1, 2, 3}",
        "d": "list = <1, 2, 3>",
        "ans": "B"
    },
    {
        "q": "What is the correct way to import a module in Python?",
        "a": "include module_name",
        "b": "require module_name",
        "c": "import module_name",
        "d": "using module_name",
        "ans": "C"
    },
    {
        "q": "What is the output of: print('abc'.upper())?",
        "a": "abc",
        "b": "Abc",
        "c": "ABC",
        "d": "Error",
        "ans": "C"
    },
    {
        "q": "Which of the following is an immutable data type?",
        "a": "List",
        "b": "Dictionary",
        "c": "Set",
        "d": "Tuple",
        "ans": "D"
    },
    {
        "q": "What does 'elif' stand for in Python?",
        "a": "else line",
        "b": "else if",
        "c": "end line",
        "d": "elif statement",
        "ans": "B"
    },
    {
        "q": "What is the output of: print(3 * 'Ha')?",
        "a": "HaHaHa",
        "b": "3Ha",
        "c": "Ha3",
        "d": "Error",
        "ans": "A"
    },
    {
        "q": "How do you start a function definition in Python?",
        "a": "function myFunc():",
        "b": "def myFunc():",
        "c": "func myFunc():",
        "d": "define myFunc():",
        "ans": "B"
    },
    {
        "q": "What is the output of: print(type([]))?",
        "a": "<class 'tuple'>",
        "b": "<class 'list'>",
        "c": "<class 'dict'>",
        "d": "<class 'array'>",
        "ans": "B"
    },
    {
        "q": "Which method removes the last element from a list?",
        "a": "remove()",
        "b": "delete()",
        "c": "pop()",
        "d": "discard()",
        "ans": "C"
    },
    {
        "q": "What is the output of: print(5 == 5.0)?",
        "a": "False",
        "b": "True",
        "c": "Error",
        "d": "None",
        "ans": "B"
    },
    {
        "q": "Which symbol is used for single-line comments in Python?",
        "a": "//",
        "b": "#",
        "c": "/*",
        "d": "--",
        "ans": "B"
    },
    {
        "q": "What is the output of: print(10 % 3)?",
        "a": "3",
        "b": "1",
        "c": "0",
        "d": "3.33",
        "ans": "B"
    },
    {
        "q": "How do you create a dictionary in Python?",
        "a": "dict = [1: 'a', 2: 'b']",
        "b": "dict = (1: 'a', 2: 'b')",
        "c": "dict = {1: 'a', 2: 'b'}",
        "d": "dict = <1: 'a', 2: 'b'>",
        "ans": "C"
    },
    {
        "q": "What does the 'break' statement do?",
        "a": "Skips the current iteration",
        "b": "Terminates the loop",
        "c": "Restarts the loop",
        "d": "Pauses the loop",
        "ans": "B"
    },
    {
        "q": "What is the output of: print('hello'.replace('h', 'H'))?",
        "a": "Hello",
        "b": "hello",
        "c": "hEllo",
        "d": "Error",
        "ans": "A"
    },
    {
        "q": "Which of the following is a valid variable name?",
        "a": "2name",
        "b": "my-name",
        "c": "_name",
        "d": "my name",
        "ans": "C"
    },
    {
        "q": "What is the output of: print(int(9.8))?",
        "a": "9.8",
        "b": "10",
        "c": "9",
        "d": "Error",
        "ans": "C"
    },
    {
        "q": "What is the correct syntax for a 'for' loop that prints numbers 0-4?",
        "a": "for i in range(4): print(i)",
        "b": "for i in range(5): print(i)",
        "c": "for i in range(1, 5): print(i)",
        "d": "for i in range(0, 4): print(i)",
        "ans": "B"
    },
    {
        "q": "What does 'pass' do in Python?",
        "a": "Exits the program",
        "b": "Does nothing (null operation)",
        "c": "Skips one iteration",
        "d": "Prints 'pass'",
        "ans": "B"
    },
    {
        "q": "What is the output of: print('Python'[0])?",
        "a": "P",
        "b": "y",
        "c": "Python",
        "d": "Error",
        "ans": "A"
    },
    {
        "q": "Which method converts a string to lowercase?",
        "a": "toLower()",
        "b": "lower()",
        "c": "downcase()",
        "d": "small()",
        "ans": "B"
    },
    {
        "q": "What is the output of: print(1 == 1 and 2 == 2)?",
        "a": "False",
        "b": "True",
        "c": "1",
        "d": "Error",
        "ans": "B"
    },
    {
        "q": "How do you create a tuple in Python?",
        "a": "tuple = [1, 2, 3]",
        "b": "tuple = (1, 2, 3)",
        "c": "tuple = {1, 2, 3}",
        "d": "tuple = <1, 2, 3>",
        "ans": "B"
    },
    {
        "q": "What is the output of: print(bool(0))?",
        "a": "True",
        "b": "1",
        "c": "False",
        "d": "None",
        "ans": "C"
    },
    {
        "q": "Which of the following is used to handle exceptions?",
        "a": "try/except",
        "b": "catch/throw",
        "c": "error/handle",
        "d": "check/resolve",
        "ans": "A"
    },
]

# --- Styles ---
title_font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

option_font = Font(name="Arial", size=11)
option_fill_a = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
option_fill_b = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
option_fill_c = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
option_fill_d = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

ans_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
ans_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

# --- Column widths ---
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 25
ws.column_dimensions["E"].width = 25
ws.column_dimensions["F"].width = 25
ws.column_dimensions["G"].width = 10

# --- Title row ---
ws.merge_cells("A1:G1")
title_cell = ws["A1"]
title_cell.value = "PYTHON OBJECTIVE QUESTIONS (MCQ) - 40 Questions"
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = center_align
title_cell.border = thin_border
ws.row_dimensions[1].height = 35

# --- Subtitle ---
ws.merge_cells("A2:G2")
sub_cell = ws["A2"]
sub_cell.value = "Choose the correct answer for each question. Each question has 4 options (A, B, C, D)."
sub_cell.font = Font(name="Arial", italic=True, size=10, color="2F5496")
sub_cell.alignment = center_align
ws.row_dimensions[2].height = 25

# --- Headers ---
headers = ["No.", "Question", "Option A", "Option B", "Option C", "Option D", "Answer"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
ws.row_dimensions[3].height = 25

# --- Questions ---
for i, q in enumerate(questions, 1):
    row = i + 3
    ws.row_dimensions[row].height = 40

    # No.
    cell = ws.cell(row=row, column=1, value=i)
    cell.font = Font(name="Arial", bold=True, size=11)
    cell.alignment = center_align
    cell.border = thin_border

    # Question
    cell = ws.cell(row=row, column=2, value=q["q"])
    cell.font = option_font
    cell.alignment = left_align
    cell.border = thin_border

    # Option A
    cell = ws.cell(row=row, column=3, value=f"A. {q['a']}")
    cell.font = option_font
    cell.fill = option_fill_a
    cell.alignment = center_align
    cell.border = thin_border

    # Option B
    cell = ws.cell(row=row, column=4, value=f"B. {q['b']}")
    cell.font = option_font
    cell.fill = option_fill_b
    cell.alignment = center_align
    cell.border = thin_border

    # Option C
    cell = ws.cell(row=row, column=5, value=f"C. {q['c']}")
    cell.font = option_font
    cell.fill = option_fill_c
    cell.alignment = center_align
    cell.border = thin_border

    # Option D
    cell = ws.cell(row=row, column=6, value=f"D. {q['d']}")
    cell.font = option_font
    cell.fill = option_fill_d
    cell.alignment = center_align
    cell.border = thin_border

    # Answer
    cell = ws.cell(row=row, column=7, value=q["ans"])
    cell.font = ans_font
    cell.fill = ans_fill
    cell.alignment = center_align
    cell.border = thin_border

# --- Instructions sheet ---
ws2 = wb.create_sheet("Instructions")
ws2.column_dimensions["A"].width = 60

instructions = [
    ("QUIZ INSTRUCTIONS", title_font, title_fill),
    ("", None, None),
    ("Subject: Python Programming (Beginner Level)", Font(name="Arial", bold=True, size=12), None),
    ("Total Questions: 40", Font(name="Arial", size=11), None),
    ("Question Type: Multiple Choice (Objective)", Font(name="Arial", size=11), None),
    ("Options per Question: 4 (A, B, C, D)", Font(name="Arial", size=11), None),
    ("", None, None),
    ("Instructions:", Font(name="Arial", bold=True, size=12), None),
    ("1. Read each question carefully before answering.", option_font, None),
    ("2. Choose only ONE correct answer from the four options.", option_font, None),
    ("3. Write the letter (A, B, C, or D) of your chosen answer.", option_font, None),
    ("4. Each question carries equal marks.", option_font, None),
    ("5. Do not use any external resources during the quiz.", option_font, None),
    ("", None, None),
    ("Topics Covered:", Font(name="Arial", bold=True, size=12), None),
    ("- Python basics and syntax", option_font, None),
    ("- Variables and data types", option_font, None),
    ("- Operators (arithmetic, comparison, logical)", option_font, None),
    ("- Strings and string methods", option_font, None),
    ("- Lists, tuples, dictionaries, and sets", option_font, None),
    ("- Control flow (if/elif/else, for, while)", option_font, None),
    ("- Functions and modules", option_font, None),
    ("- Exception handling (try/except)", option_font, None),
    ("- Type casting and boolean values", option_font, None),
]

for i, (text, font, fill) in enumerate(instructions, 1):
    cell = ws2.cell(row=i, column=1, value=text)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center")

# Save
filepath = "/home/viktr/viktr_dfi/digitafort/py_be/Python_MCQ_Quiz_40_Questions.xlsx"
wb.save(filepath)
print(f"Excel file saved: {filepath}")
print(f"Total questions: {len(questions)}")
